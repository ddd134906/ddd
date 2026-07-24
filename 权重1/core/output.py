import pandas as pd
import numpy as np
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule


def save_results(output_path, desc_list, loadings, score_coeff,
                 factor_corr, reg_results, qd_stats=None, qd_desc=None,
                 qd_means=None, channel_means_dict=None, diff_table=None,
                 country='', fa_extra=None, kpi_check_df=None,
                 channel_check_df=None, sample_check_success=False):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        write_descriptives(writer, desc_list)
        if kpi_check_df is not None and channel_check_df is not None:
            write_sample_check(writer, kpi_check_df, channel_check_df, sample_check_success)
        if fa_extra is not None:
            write_cfa_results(writer, fa_extra)
        if factor_corr is not None:
            write_factor_corr(writer, factor_corr, fa_extra)
        write_reg_results(writer, reg_results)
        if qd_stats is not None and qd_means is not None:
            write_seen_vs_noseen(writer, qd_stats, qd_desc, qd_means,
                                 channel_means_dict, diff_table)

# ========== 描述统计 ==========
def write_descriptives(writer, desc_list):
    ws = writer.book.create_sheet('Descriptives')
    current_row = 1

    def write_num(cell, val):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0' if val == int(val) else '0.00'

    for title, df_desc in desc_list:
        if df_desc is not None and not df_desc.empty:
            ws.cell(row=current_row, column=1, value=title)
            current_row += 1
            headers = ['Variable', 'N', 'Min', 'Max', 'Mean', 'Std']
            for col, h in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col, value=h)
            current_row += 1
            for idx, row in df_desc.iterrows():
                ws.cell(row=current_row, column=1, value=idx)
                for col, field in enumerate(['N', 'Min', 'Max', 'Mean', 'Std'], start=2):
                    write_num(ws.cell(row=current_row, column=col), row[field])
                current_row += 1
            current_row += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)


# ========== 样本量校对 ==========
def write_sample_check(writer, kpi_check_df, channel_check_df, success):
    ws = writer.book.create_sheet('Sample_Check')
    row = 1

    # KPI 表
    ws.cell(row=row, column=1, value='KPI 样本量校对')
    row += 1
    headers = ['表名', 'KPI标签', 'Full_Table样本量', 'Mean*N (描述统计)', '差异']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1
    for _, row_data in kpi_check_df.iterrows():
        ws.cell(row=row, column=1, value=row_data['TableName'])
        ws.cell(row=row, column=2, value=row_data['KPI_Label'])
        ws.cell(row=row, column=3, value=row_data['FullTable_Sample'])
        ws.cell(row=row, column=4, value=row_data['Calc_Mean*N'])
        ws.cell(row=row, column=5, value=row_data['Diff'])
        row += 1

    row += 1

    # 渠道表
    ws.cell(row=row, column=1, value='渠道样本量校对')
    row += 1
    headers = ['表名', '渠道名', 'Full_Table样本量', 'Mean*N (描述统计)', '差异']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1
    for _, row_data in channel_check_df.iterrows():
        ws.cell(row=row, column=1, value=row_data['TableName'])
        ws.cell(row=row, column=2, value=row_data['Channel'])
        ws.cell(row=row, column=3, value=row_data['FullTable_Sample'])
        ws.cell(row=row, column=4, value=row_data['Calc_Mean*N'])
        ws.cell(row=row, column=5, value=row_data['Diff'])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='样本量校对结果：')
    ws.cell(row=row, column=2, value='成功' if success else '失败（存在差异）')

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)


# ========== FactorCorr ==========
def write_factor_corr(writer, factor_corr, fa_extra):
    ws = writer.book.create_sheet('FactorCorr')
    n_factors = fa_extra.get('n_factors', factor_corr.shape[0])
    factor_labels = [f'Factor{i+1}' for i in range(n_factors)]

    df_corr = pd.DataFrame(factor_corr, index=factor_labels, columns=factor_labels)
    df_corr.to_excel(writer, sheet_name='FactorCorr', startrow=0, startcol=0)

    ws = writer.sheets['FactorCorr']
    for r in range(1, n_factors + 1):
        for c in range(1, n_factors + 1):
            cell = ws.cell(row=r + 1, column=c + 1)
            val = cell.value
            if val is not None:
                if abs(val) < 1e-9:
                    cell.number_format = '0.000'
                else:
                    if abs(val - round(val)) < 1e-12:
                        cell.number_format = '0'
                    else:
                        cell.number_format = '0.000'

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 12)


# ========== CFA 结果 ==========
def write_cfa_results(writer, fa_extra):
    ws = writer.book.create_sheet('CFA_Results')
    row = 1
    col_shift = 1

    def write_num(cell, val, fmt='0.00'):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0' if val == int(val) else fmt

    # ---- 头部信息 ----
    ws.cell(row=row, column=col_shift, value="Spend time = 00:00:02")
    row += 1
    ws.cell(row=row, column=col_shift, value="Factor Result")
    row += 1
    now = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    ws.cell(row=row, column=col_shift, value=f"TIME= {now}")
    row += 1
    rot_method = fa_extra.get('rotation', 'varimax').capitalize()
    ws.cell(row=row, column=col_shift, value=f"Rotation Method= {rot_method} Method")
    row += 3

    # ---- 1. Total Variance Explained ----
    ws.cell(row=row, column=col_shift, value="Total Variance Explained")
    row += 1
    ws.cell(row=row, column=col_shift, value="Initial Eigenvalues")
    row += 1
    ws.cell(row=row, column=col_shift, value="Extraction Method: Principal Component Analysis.")
    row += 2

    eigenvals = fa_extra.get('eigenvals')
    var_ratio = fa_extra.get('var_ratio')
    cum_var = fa_extra.get('cum_var')
    if eigenvals is not None:
        ws.cell(row=row, column=col_shift, value="Component")
        ws.cell(row=row, column=col_shift+1, value="Total")
        ws.cell(row=row, column=col_shift+2, value="% of Variance")
        ws.cell(row=row, column=col_shift+3, value="Cumulative %")
        row += 1
        for i, (ev, vr, cv) in enumerate(zip(eigenvals, var_ratio, cum_var), start=1):
            ws.cell(row=row, column=col_shift, value=f"Factor{i}")
            write_num(ws.cell(row=row, column=col_shift+1), ev)
            write_num(ws.cell(row=row, column=col_shift+2), vr, fmt='0.00%')
            write_num(ws.cell(row=row, column=col_shift+3), cv, fmt='0.00%')
            row += 1
        row += 1

    # ---- 获取基础数据 ----
    target = fa_extra.get('target')
    if target is None:
        raise ValueError("fa_extra 中缺少 'target'")
    row_labels = target.index.tolist()
    col_labels = target.columns.tolist()
    n_rows = len(row_labels)
    n_factors = len(col_labels)
    sig_break = fa_extra.get('sig_break', 0.3)

    # ---- 2. Component Matrix（未旋转，原始顺序） ----
    unrot = fa_extra.get('unrot_loadings')
    if unrot is not None:
        if unrot.shape[0] != n_rows or unrot.shape[1] != n_factors:
            raise ValueError(f"unrot_loadings 形状 {unrot.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Component Matrix")
        row += 1
        ws.cell(row=row, column=col_shift, value="Extraction Method: Principal Component Analysis.")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)

        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                cell = ws.cell(row=row, column=col_shift+3+j)
                write_num(cell, unrot[i, j])
                if abs(unrot[i, j]) > sig_break:
                    cell.fill = yellow_fill
                    cell.font = bold_font
            row += 1
        row += 1

    # ---- 3. Rotated Component Matrix（旋转载荷，原始顺序） ----
    rot = fa_extra.get('rot_loadings')
    if rot is not None:
        if rot.shape[0] != n_rows or rot.shape[1] != n_factors:
            raise ValueError(f"rot_loadings 形状 {rot.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Rotated Component Matrix")
        row += 1
        ws.cell(row=row, column=col_shift, value="Rotation Method: Procrustes with Kaiser Normalization.")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                cell = ws.cell(row=row, column=col_shift+3+j)
                write_num(cell, rot[i, j])
                if target.iloc[i, j] != 0:
                    cell.font = Font(italic=True, color="FF0000")
                if abs(rot[i, j]) > sig_break:
                    if target.iloc[i, j] != 0:
                        cell.font = Font(bold=True, italic=True, color="FF0000")
                    else:
                        cell.font = Font(bold=True)
                    cell.fill = yellow_fill
            row += 1
        row += 1

    # ---- 4. Available Rotation Rule Matrix ----
    ws.cell(row=row, column=col_shift, value="Available Rotation Rule Matrix")
    row += 1
    ws.cell(row=row, column=col_shift, value="You can Copy this Matrix to Rotation Rule Sheet.")
    row += 2
    ws.cell(row=row, column=col_shift, value="VariableID")
    ws.cell(row=row, column=col_shift+1, value="Variable")
    ws.cell(row=row, column=col_shift+2, value="Lable")
    for j, col_name in enumerate(col_labels):
        ws.cell(row=row, column=col_shift+3+j, value=col_name)
    row += 1

    for i, var_name in enumerate(row_labels):
        ws.cell(row=row, column=col_shift, value=i+1)
        ws.cell(row=row, column=col_shift+1, value=var_name)
        ws.cell(row=row, column=col_shift+2, value=var_name)
        for j in range(n_factors):
            val = rot[i, j]
            if val > sig_break:
                ws.cell(row=row, column=col_shift+3+j, value=1)
            elif val < -sig_break:
                ws.cell(row=row, column=col_shift+3+j, value=-1)
            else:
                ws.cell(row=row, column=col_shift+3+j, value='')
        row += 1
    row += 1

    # ---- 5. Component Score Coefficient Matrix（标准化，原始顺序） ----
    score_std = fa_extra.get('score_coeff_std')
    if score_std is not None:
        if score_std.shape[0] != n_rows or score_std.shape[1] != n_factors:
            raise ValueError(f"score_coeff_std 形状 {score_std.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Component Score Coefficient Matrix")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                write_num(ws.cell(row=row, column=col_shift+3+j), score_std[i, j])
            row += 1
        row += 1

    # ---- 6. Unstandardized Component Score Coefficient Matrix（使用表格版本） ----
    score_unstd_table = fa_extra.get('score_coeff_unstd_table')
    if score_unstd_table is not None:
        if score_unstd_table.shape[0] != n_rows or score_unstd_table.shape[1] != n_factors:
            raise ValueError(f"score_coeff_unstd_table 形状 {score_unstd_table.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Unstandarized Component Score Coefficient Matrix")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                write_num(ws.cell(row=row, column=col_shift+3+j), score_unstd_table[i, j])
            row += 1
        row += 1

    # ---- 7. SPSS Syntax ----
    spss_syntax = fa_extra.get('spss_syntax')
    if spss_syntax:
        ws.cell(row=row, column=col_shift, value="SPSS Syntax")
        row += 2
        for line in spss_syntax.split('\n'):
            ws.cell(row=row, column=col_shift, value=line)
            row += 1

    # ---- 8. 自动列宽 ----
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)


# ========== 回归结果 ==========
def write_reg_results(writer, reg_results):
    ws = writer.book.create_sheet('reg_Results')
    current_row = 1

    def write_num(cell, val):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0' if val == int(val) else '0.00'

    all_models = []
    for model_name, res_dict in reg_results.items():
        for dep, res in res_dict.items():
            all_models.append((dep, res))
    if not all_models:
        return

    model_counter = 1
    for dep, res in all_models:
        r2 = res.get('R_squared', np.nan)
        adj_r2 = res.get('adj_R_squared', np.nan)
        r = res.get('R', np.nan)
        std_err = res.get('std_error', np.nan)
        coeff = res.get('coeff', pd.Series())
        bse = res.get('bse', pd.Series())
        beta = res.get('beta', pd.Series())
        tvals = res.get('tvalues', pd.Series())
        pvals = res.get('pvalues', pd.Series())
        ci_lower = res.get('ci_lower', pd.Series())
        ci_upper = res.get('ci_upper', pd.Series())
        var_names = res.get('var_names', list(coeff.index))
        if not var_names:
            var_names = list(coeff.index)

        ws.cell(row=current_row, column=1, value=f'Dependent Variable: {dep}')
        current_row += 1

        ws.cell(row=current_row, column=1, value='Model Summary')
        current_row += 1
        ws.cell(row=current_row, column=1, value='Model')
        ws.cell(row=current_row, column=2, value='R')
        ws.cell(row=current_row, column=3, value='R Square')
        ws.cell(row=current_row, column=4, value='Adjusted R Square')
        ws.cell(row=current_row, column=5, value='Std. Error of the Estimate')
        current_row += 1
        ws.cell(row=current_row, column=1, value=model_counter)
        write_num(ws.cell(row=current_row, column=2), r)
        write_num(ws.cell(row=current_row, column=3), r2)
        write_num(ws.cell(row=current_row, column=4), adj_r2)
        write_num(ws.cell(row=current_row, column=5), std_err)
        current_row += 1

        predictors = [v for v in var_names if v != 'const']
        pred_str = '(Constant), ' + ', '.join(predictors)
        ws.cell(row=current_row, column=1, value=f'a. Predictors: {pred_str}')
        current_row += 2

        ws.cell(row=current_row, column=1, value='Coefficientsa')
        current_row += 1
        ws.cell(row=current_row, column=1, value='Model')
        ws.cell(row=current_row, column=2, value='')
        ws.cell(row=current_row, column=3, value='Unstandardized Coefficients')
        ws.cell(row=current_row, column=4, value='')
        ws.cell(row=current_row, column=5, value='Standardized Coefficients')
        ws.cell(row=current_row, column=6, value='t')
        ws.cell(row=current_row, column=7, value='Sig.')
        ws.cell(row=current_row, column=8, value='95.0% Confidence Interval for B')
        ws.cell(row=current_row, column=9, value='')
        current_row += 1
        ws.cell(row=current_row, column=1, value='')
        ws.cell(row=current_row, column=2, value='')
        ws.cell(row=current_row, column=3, value='B')
        ws.cell(row=current_row, column=4, value='Std. Error')
        ws.cell(row=current_row, column=5, value='Beta')
        ws.cell(row=current_row, column=6, value='')
        ws.cell(row=current_row, column=7, value='')
        ws.cell(row=current_row, column=8, value='Lower Bound')
        ws.cell(row=current_row, column=9, value='Upper Bound')
        current_row += 1

        for i, var in enumerate(var_names):
            if i == 0:
                ws.cell(row=current_row, column=1, value=model_counter)
            else:
                ws.cell(row=current_row, column=1, value='')
            # 仅将 'const' 显示为 'constant'
            display_name = 'constant' if var == 'const' else var
            ws.cell(row=current_row, column=2, value=display_name)
            write_num(ws.cell(row=current_row, column=3), coeff[var])
            write_num(ws.cell(row=current_row, column=4), bse[var])
            beta_val = beta.get(var, np.nan)
            ws.cell(row=current_row, column=5, value=beta_val if not np.isnan(beta_val) else '')
            write_num(ws.cell(row=current_row, column=6), tvals[var])
            write_num(ws.cell(row=current_row, column=7), pvals[var])
            write_num(ws.cell(row=current_row, column=8), ci_lower[var])
            write_num(ws.cell(row=current_row, column=9), ci_upper[var])
            current_row += 1

        ws.cell(row=current_row, column=2, value=f'a. Dependent Variable: {dep}')
        current_row += 2
        model_counter += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)


# ========== Seen vs No Seen ==========
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

def write_seen_vs_noseen(writer, qd_stats, qd_desc, qd_means, channel_means_dict, diff_table):
    """
    将三张表写入同一个工作表 'seen vs no seen'：
    1. qd 描述统计（N, Min, Max, Mean, Std）
    2. 均值对比表（qd=0/1，各渠道 No/Yes）
    3. 差值表（qd (Diff)，各渠道 (Diff)），数值显示为百分比，
       并添加与截图完全一致的三色标尺条件格式（红-黄-绿）
    """
    ws = writer.book.create_sheet('seen vs no seen')
    current_row = 1

    def write_num(cell, val):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0.00' if val != int(val) else '0'

    def write_percent(cell, val):
        """写入百分比数值，格式为 0.00%"""
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0.00%'

    # ---------- 1. qd 描述统计 ----------
    ws.cell(row=current_row, column=1, value='Descriptive Statistics')
    current_row += 1
    headers = ['', 'N', 'Minimum', 'Maximum', 'Mean', 'Std. Deviation']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=current_row, column=col, value=h)
    current_row += 1

    if not qd_stats.empty:
        row_data = qd_stats.iloc[0]
        ws.cell(row=current_row, column=1, value='qd')
        write_num(ws.cell(row=current_row, column=2), row_data['count'])
        write_num(ws.cell(row=current_row, column=3), row_data['min'])
        write_num(ws.cell(row=current_row, column=4), row_data['max'])
        write_num(ws.cell(row=current_row, column=5), row_data['mean'])
        write_num(ws.cell(row=current_row, column=6), row_data['std'])
    current_row += 2

    # ---------- 2. 均值对比表 ----------
    channel_names = list(channel_means_dict.keys()) if channel_means_dict else []
    mean_headers = ['']
    if not qd_means.empty and 0 in qd_means.columns and 1 in qd_means.columns:
        mean_headers.append('qd=0')
        mean_headers.append('qd=1')
    else:
        mean_headers.append('qd=0')
        mean_headers.append('qd=1')
    for ch in channel_names:
        mean_headers.append(f'{ch} (No)')
        mean_headers.append(f'{ch} (Yes)')

    for col, h in enumerate(mean_headers, start=1):
        ws.cell(row=current_row, column=col, value=h)
    current_row += 1

    y_vars = qd_means.index.tolist() if not qd_means.empty else diff_table.index.tolist()
    for y_var in y_vars:
        ws.cell(row=current_row, column=1, value=y_var)
        col = 2
        if not qd_means.empty and y_var in qd_means.index:
            write_num(ws.cell(row=current_row, column=col), qd_means.loc[y_var, 0])
            write_num(ws.cell(row=current_row, column=col+1), qd_means.loc[y_var, 1])
        else:
            ws.cell(row=current_row, column=col, value='')
            ws.cell(row=current_row, column=col+1, value='')
        col += 2
        for ch in channel_names:
            gm = channel_means_dict.get(ch)
            if gm is not None and y_var in gm.columns:
                write_num(ws.cell(row=current_row, column=col), gm.loc[0, y_var])
                write_num(ws.cell(row=current_row, column=col+1), gm.loc[1, y_var])
            else:
                ws.cell(row=current_row, column=col, value='')
                ws.cell(row=current_row, column=col+1, value='')
            col += 2
        current_row += 1

    current_row += 1

    # ---------- 3. 差值表 ----------
    diff_start_row = current_row
    diff_headers = [''] + list(diff_table.columns)
    for col, h in enumerate(diff_headers, start=1):
        ws.cell(row=current_row, column=col, value=h)
    current_row += 1

    # 写入差值数据，使用百分比格式
    for y_var in y_vars:
        ws.cell(row=current_row, column=1, value=y_var)
        for col_idx, col_name in enumerate(diff_table.columns, start=2):
            if y_var in diff_table.index:
                val = diff_table.loc[y_var, col_name]
                write_percent(ws.cell(row=current_row, column=col_idx), val)
            else:
                ws.cell(row=current_row, column=col_idx, value='')
        current_row += 1

    diff_end_row = current_row - 1

    # ---------- 添加条件格式（红-黄-绿三色标尺，与截图完全一致） ----------
    if not diff_table.empty and len(diff_table.columns) > 0:
        for col_idx in range(2, 2 + len(diff_table.columns)):
            col_letter = get_column_letter(col_idx)
            range_str = f'{col_letter}{diff_start_row + 1}:{col_letter}{diff_end_row}'
            # 三色标尺：最低值→红，百分位数50→黄，最高值→绿
            rule = ColorScaleRule(
                start_type='min', start_color='F8696B',      # 红色
                mid_type='percentile', mid_value=50, mid_color='FFEB84',  # 黄色
                end_type='max', end_color='63BE7B'          # 绿色
            )
            ws.conditional_formatting.add(range_str, rule)

    # ---------- 自动列宽（优化版：数值列宽度10，文本列动态但不超过20） ----------
    for col in ws.columns:
        col_idx = col[0].column
        max_len = 0
        for cell in col:
            if cell.value is not None:
                # 对百分比格式的单元格，预估显示长度（如 -12.34% 为7个字符）
                if isinstance(cell.value, float) and cell.number_format == '0.00%':
                    max_len = max(max_len, 7)
                else:
                    max_len = max(max_len, len(str(cell.value)))
        if col_idx >= 2:
            width = 10   # 数值列固定宽度（足够显示 -12.34%）
        else:
            width = min(max_len + 2, 20)  # 文本列动态但不超过20
        ws.column_dimensions[get_column_letter(col_idx)].width = width