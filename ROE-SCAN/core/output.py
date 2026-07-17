import pandas as pd
import numpy as np
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font


def save_results(output_path, desc_list, loadings, score_coeff,
                 factor_corr, reg_results, qd_means, channel_means,
                 country='', fa_extra=None, kpi_check_df=None,
                 channel_check_df=None, sample_check_success=False):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 1. Descriptives（描述统计）
        write_descriptives(writer, desc_list)

        # 2. Sample_Check（样本量校对）
        if kpi_check_df is not None and channel_check_df is not None:
            write_sample_check(writer, kpi_check_df, channel_check_df, sample_check_success)

        # 3. CFA_Results
        if fa_extra is not None:
            write_cfa_results(writer, fa_extra)

        # 4. FactorCorr
        if factor_corr is not None:
            write_factor_corr(writer, factor_corr, fa_extra)

        # 5. reg_Results
        write_reg_results(writer, reg_results)

        # 6. seen_vs_no_seen
        write_seen_vs_noseen(writer, qd_means, channel_means)


# ========== 描述统计 ==========
def write_descriptives(writer, desc_list):
    ws = writer.book.create_sheet('Descriptives')
    current_row = 1

    def write_num(cell, val):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0' if val == int(val) else '0.00'

    for title, df_desc in desc_list:
        if df_desc is not None and not df_desc.empty:
            ws.cell(row=current_row, column=1, value=title)
            current_row += 1
            headers = ['Variable', 'N', 'Min', 'Max', 'Mean', 'Std']
            for col, h in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col, value=h)
            current_row += 1
            for idx, row in df_desc.iterrows():
                ws.cell(row=current_row, column=1, value=idx)
                for col, field in enumerate(['N', 'Min', 'Max', 'Mean', 'Std'], start=2):
                    write_num(ws.cell(row=current_row, column=col), row[field])
                current_row += 1
            current_row += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)


# ========== 样本量校对 ==========
def write_sample_check(writer, kpi_check_df, channel_check_df, success):
    ws = writer.book.create_sheet('Sample_Check')
    row = 1

    # KPI 表
    ws.cell(row=row, column=1, value='KPI 样本量校对')
    row += 1
    headers = ['表名', 'KPI标签', 'Full_Table样本量', 'Mean*N (描述统计)', '差异']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1
    for _, row_data in kpi_check_df.iterrows():
        ws.cell(row=row, column=1, value=row_data['TableName'])
        ws.cell(row=row, column=2, value=row_data['KPI_Label'])
        ws.cell(row=row, column=3, value=row_data['FullTable_Sample'])
        ws.cell(row=row, column=4, value=row_data['Calc_Mean*N'])
        ws.cell(row=row, column=5, value=row_data['Diff'])
        row += 1

    row += 1

    # 渠道表
    ws.cell(row=row, column=1, value='渠道样本量校对')
    row += 1
    headers = ['表名', '渠道名', 'Full_Table样本量', 'Mean*N (描述统计)', '差异']
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1
    for _, row_data in channel_check_df.iterrows():
        ws.cell(row=row, column=1, value=row_data['TableName'])
        ws.cell(row=row, column=2, value=row_data['Channel'])
        ws.cell(row=row, column=3, value=row_data['FullTable_Sample'])
        ws.cell(row=row, column=4, value=row_data['Calc_Mean*N'])
        ws.cell(row=row, column=5, value=row_data['Diff'])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='样本量校对结果：')
    ws.cell(row=row, column=2, value='成功' if success else '失败（存在差异）')

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)


# ========== FactorCorr ==========
def write_factor_corr(writer, factor_corr, fa_extra):
    ws = writer.book.create_sheet('FactorCorr')
    n_factors = fa_extra.get('n_factors', factor_corr.shape[0])
    factor_labels = [f'Factor{i+1}' for i in range(n_factors)]

    df_corr = pd.DataFrame(factor_corr, index=factor_labels, columns=factor_labels)
    df_corr.to_excel(writer, sheet_name='FactorCorr', startrow=0, startcol=0)

    ws = writer.sheets['FactorCorr']
    for r in range(1, n_factors + 1):
        for c in range(1, n_factors + 1):
            cell = ws.cell(row=r + 1, column=c + 1)
            val = cell.value
            if val is not None:
                if abs(val) < 1e-9:
                    cell.number_format = '0.000'
                else:
                    if abs(val - round(val)) < 1e-12:
                        cell.number_format = '0'
                    else:
                        cell.number_format = '0.000'

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 12)


# ========== CFA 结果 ==========
def write_cfa_results(writer, fa_extra):
    ws = writer.book.create_sheet('CFA_Results')
    row = 1
    col_shift = 1

    def write_num(cell, val, fmt='0.00'):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0' if val == int(val) else fmt

    # ---- 头部信息 ----
    ws.cell(row=row, column=col_shift, value="Spend time = 00:00:02")
    row += 1
    ws.cell(row=row, column=col_shift, value="Factor Result")
    row += 1
    now = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    ws.cell(row=row, column=col_shift, value=f"TIME= {now}")
    row += 1
    rot_method = fa_extra.get('rotation', 'varimax').capitalize()
    ws.cell(row=row, column=col_shift, value=f"Rotation Method= {rot_method} Method")
    row += 3

    # ---- 1. Total Variance Explained ----
    ws.cell(row=row, column=col_shift, value="Total Variance Explained")
    row += 1
    ws.cell(row=row, column=col_shift, value="Initial Eigenvalues")
    row += 1
    ws.cell(row=row, column=col_shift, value="Extraction Method: Principal Component Analysis.")
    row += 2

    eigenvals = fa_extra.get('eigenvals')
    var_ratio = fa_extra.get('var_ratio')
    cum_var = fa_extra.get('cum_var')
    if eigenvals is not None:
        ws.cell(row=row, column=col_shift, value="Component")
        ws.cell(row=row, column=col_shift+1, value="Total")
        ws.cell(row=row, column=col_shift+2, value="% of Variance")
        ws.cell(row=row, column=col_shift+3, value="Cumulative %")
        row += 1
        for i, (ev, vr, cv) in enumerate(zip(eigenvals, var_ratio, cum_var), start=1):
            ws.cell(row=row, column=col_shift, value=f"Factor{i}")
            write_num(ws.cell(row=row, column=col_shift+1), ev)
            write_num(ws.cell(row=row, column=col_shift+2), vr, fmt='0.00%')
            write_num(ws.cell(row=row, column=col_shift+3), cv, fmt='0.00%')
            row += 1
        row += 1

    # ---- 获取基础数据 ----
    target = fa_extra.get('target')
    if target is None:
        raise ValueError("fa_extra 中缺少 'target'")
    row_labels = target.index.tolist()
    col_labels = target.columns.tolist()
    n_rows = len(row_labels)
    n_factors = len(col_labels)
    sig_break = fa_extra.get('sig_break', 0.3)

    # ---- 2. Component Matrix（未旋转，原始顺序） ----
    unrot = fa_extra.get('unrot_loadings')
    if unrot is not None:
        if unrot.shape[0] != n_rows or unrot.shape[1] != n_factors:
            raise ValueError(f"unrot_loadings 形状 {unrot.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Component Matrix")
        row += 1
        ws.cell(row=row, column=col_shift, value="Extraction Method: Principal Component Analysis.")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)

        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                cell = ws.cell(row=row, column=col_shift+3+j)
                write_num(cell, unrot[i, j])
                if abs(unrot[i, j]) > sig_break:
                    cell.fill = yellow_fill
                    cell.font = bold_font
            row += 1
        row += 1

    # ---- 3. Rotated Component Matrix（旋转载荷，原始顺序） ----
    rot = fa_extra.get('rot_loadings')
    if rot is not None:
        if rot.shape[0] != n_rows or rot.shape[1] != n_factors:
            raise ValueError(f"rot_loadings 形状 {rot.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Rotated Component Matrix")
        row += 1
        ws.cell(row=row, column=col_shift, value="Rotation Method: Procrustes with Kaiser Normalization.")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                cell = ws.cell(row=row, column=col_shift+3+j)
                write_num(cell, rot[i, j])
                if target.iloc[i, j] != 0:
                    cell.font = Font(italic=True, color="FF0000")
                if abs(rot[i, j]) > sig_break:
                    if target.iloc[i, j] != 0:
                        cell.font = Font(bold=True, italic=True, color="FF0000")
                    else:
                        cell.font = Font(bold=True)
                    cell.fill = yellow_fill
            row += 1
        row += 1

    # ---- 4. Available Rotation Rule Matrix ----
    ws.cell(row=row, column=col_shift, value="Available Rotation Rule Matrix")
    row += 1
    ws.cell(row=row, column=col_shift, value="You can Copy this Matrix to Rotation Rule Sheet.")
    row += 2
    ws.cell(row=row, column=col_shift, value="VariableID")
    ws.cell(row=row, column=col_shift+1, value="Variable")
    ws.cell(row=row, column=col_shift+2, value="Lable")
    for j, col_name in enumerate(col_labels):
        ws.cell(row=row, column=col_shift+3+j, value=col_name)
    row += 1

    for i, var_name in enumerate(row_labels):
        ws.cell(row=row, column=col_shift, value=i+1)
        ws.cell(row=row, column=col_shift+1, value=var_name)
        ws.cell(row=row, column=col_shift+2, value=var_name)
        for j in range(n_factors):
            val = rot[i, j]
            if val > sig_break:
                ws.cell(row=row, column=col_shift+3+j, value=1)
            elif val < -sig_break:
                ws.cell(row=row, column=col_shift+3+j, value=-1)
            else:
                ws.cell(row=row, column=col_shift+3+j, value='')
        row += 1
    row += 1

    # ---- 5. Component Score Coefficient Matrix（标准化，原始顺序） ----
    score_std = fa_extra.get('score_coeff_std')
    if score_std is not None:
        if score_std.shape[0] != n_rows or score_std.shape[1] != n_factors:
            raise ValueError(f"score_coeff_std 形状 {score_std.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Component Score Coefficient Matrix")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                write_num(ws.cell(row=row, column=col_shift+3+j), score_std[i, j])
            row += 1
        row += 1

    # ---- 6. Unstandardized Component Score Coefficient Matrix（使用表格版本） ----
    score_unstd_table = fa_extra.get('score_coeff_unstd_table')
    if score_unstd_table is not None:
        if score_unstd_table.shape[0] != n_rows or score_unstd_table.shape[1] != n_factors:
            raise ValueError(f"score_coeff_unstd_table 形状 {score_unstd_table.shape} 与目标矩阵 {target.shape} 不一致")
        ws.cell(row=row, column=col_shift, value="Unstandarized Component Score Coefficient Matrix")
        row += 2
        ws.cell(row=row, column=col_shift, value="VariableID")
        ws.cell(row=row, column=col_shift+1, value="Variable")
        ws.cell(row=row, column=col_shift+2, value="Lable")
        for j, col_name in enumerate(col_labels):
            ws.cell(row=row, column=col_shift+3+j, value=col_name)
        row += 1
        for i, var_name in enumerate(row_labels):
            ws.cell(row=row, column=col_shift, value=i+1)
            ws.cell(row=row, column=col_shift+1, value=var_name)
            ws.cell(row=row, column=col_shift+2, value=var_name)
            for j in range(n_factors):
                write_num(ws.cell(row=row, column=col_shift+3+j), score_unstd_table[i, j])
            row += 1
        row += 1

    # ---- 7. SPSS Syntax ----
    spss_syntax = fa_extra.get('spss_syntax')
    if spss_syntax:
        ws.cell(row=row, column=col_shift, value="SPSS Syntax")
        row += 2
        for line in spss_syntax.split('\n'):
            ws.cell(row=row, column=col_shift, value=line)
            row += 1

    # ---- 8. 自动列宽 ----
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)


# ========== 回归结果 ==========
def write_reg_results(writer, reg_results):
    ws = writer.book.create_sheet('reg_Results')
    current_row = 1

    def write_num(cell, val):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0' if val == int(val) else '0.00'

    all_models = []
    for model_name, res_dict in reg_results.items():
        for dep, res in res_dict.items():
            all_models.append((dep, res))
    if not all_models:
        return

    model_counter = 1
    for dep, res in all_models:
        r2 = res.get('R_squared', np.nan)
        adj_r2 = res.get('adj_R_squared', np.nan)
        r = res.get('R', np.nan)
        std_err = res.get('std_error', np.nan)
        coeff = res.get('coeff', pd.Series())
        bse = res.get('bse', pd.Series())
        beta = res.get('beta', pd.Series())
        tvals = res.get('tvalues', pd.Series())
        pvals = res.get('pvalues', pd.Series())
        ci_lower = res.get('ci_lower', pd.Series())
        ci_upper = res.get('ci_upper', pd.Series())
        var_names = res.get('var_names', list(coeff.index))
        if not var_names:
            var_names = list(coeff.index)

        ws.cell(row=current_row, column=1, value=f'Dependent Variable: {dep}')
        current_row += 1

        ws.cell(row=current_row, column=1, value='Model Summary')
        current_row += 1
        ws.cell(row=current_row, column=1, value='Model')
        ws.cell(row=current_row, column=2, value='R')
        ws.cell(row=current_row, column=3, value='R Square')
        ws.cell(row=current_row, column=4, value='Adjusted R Square')
        ws.cell(row=current_row, column=5, value='Std. Error of the Estimate')
        current_row += 1
        ws.cell(row=current_row, column=1, value=model_counter)
        write_num(ws.cell(row=current_row, column=2), r)
        write_num(ws.cell(row=current_row, column=3), r2)
        write_num(ws.cell(row=current_row, column=4), adj_r2)
        write_num(ws.cell(row=current_row, column=5), std_err)
        current_row += 1

        predictors = [v for v in var_names if v != 'const']
        pred_str = '(Constant), ' + ', '.join(predictors)
        ws.cell(row=current_row, column=1, value=f'a. Predictors: {pred_str}')
        current_row += 2

        ws.cell(row=current_row, column=1, value='Coefficientsa')
        current_row += 1
        ws.cell(row=current_row, column=1, value='Model')
        ws.cell(row=current_row, column=2, value='')
        ws.cell(row=current_row, column=3, value='Unstandardized Coefficients')
        ws.cell(row=current_row, column=4, value='')
        ws.cell(row=current_row, column=5, value='Standardized Coefficients')
        ws.cell(row=current_row, column=6, value='t')
        ws.cell(row=current_row, column=7, value='Sig.')
        ws.cell(row=current_row, column=8, value='95.0% Confidence Interval for B')
        ws.cell(row=current_row, column=9, value='')
        current_row += 1
        ws.cell(row=current_row, column=1, value='')
        ws.cell(row=current_row, column=2, value='')
        ws.cell(row=current_row, column=3, value='B')
        ws.cell(row=current_row, column=4, value='Std. Error')
        ws.cell(row=current_row, column=5, value='Beta')
        ws.cell(row=current_row, column=6, value='')
        ws.cell(row=current_row, column=7, value='')
        ws.cell(row=current_row, column=8, value='Lower Bound')
        ws.cell(row=current_row, column=9, value='Upper Bound')
        current_row += 1

        for i, var in enumerate(var_names):
            if i == 0:
                ws.cell(row=current_row, column=1, value=model_counter)
            else:
                ws.cell(row=current_row, column=1, value='')
            # 仅将 'const' 显示为 'constant'
            display_name = 'constant' if var == 'const' else var
            ws.cell(row=current_row, column=2, value=display_name)
            write_num(ws.cell(row=current_row, column=3), coeff[var])
            write_num(ws.cell(row=current_row, column=4), bse[var])
            beta_val = beta.get(var, np.nan)
            ws.cell(row=current_row, column=5, value=beta_val if not np.isnan(beta_val) else '')
            write_num(ws.cell(row=current_row, column=6), tvals[var])
            write_num(ws.cell(row=current_row, column=7), pvals[var])
            write_num(ws.cell(row=current_row, column=8), ci_lower[var])
            write_num(ws.cell(row=current_row, column=9), ci_upper[var])
            current_row += 1

        ws.cell(row=current_row, column=2, value=f'a. Dependent Variable: {dep}')
        current_row += 2
        model_counter += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)


# ========== Seen vs No Seen ==========
def write_seen_vs_noseen(writer, qd_means, channel_means):
    ws = writer.book.create_sheet('seen_vs_no_seen')
    current_row = 1

    def write_num(cell, val):
        cell.value = val
        if isinstance(val, (int, float)):
            cell.number_format = '0' if val == int(val) else '0.00'

    if qd_means is not None and not qd_means.empty:
        ws.cell(row=current_row, column=1, value='Seen vs No Seen (qd grouping)')
        current_row += 1
        qd_T = qd_means.T
        for idx, row in qd_T.iterrows():
            ws.cell(row=current_row, column=1, value=idx)
            for col, val in enumerate(row, start=2):
                write_num(ws.cell(row=current_row, column=col), val)
            current_row += 1
        current_row += 1

    if channel_means is not None and not channel_means.empty:
        ws.cell(row=current_row, column=1, value='Seen vs No Seen (channel grouping)')
        current_row += 1
        for idx, row in channel_means.iterrows():
            ws.cell(row=current_row, column=1, value=idx)
            for col, val in enumerate(row, start=2):
                write_num(ws.cell(row=current_row, column=col), val)
            current_row += 1

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 20)